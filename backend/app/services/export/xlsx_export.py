"""Render the spreadsheet editor's grid into .xlsx / .csv."""
from __future__ import annotations

import csv
import io
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NUMERIC_RE = re.compile(r"^-?[\d,]*\.?\d+%?$")
HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
THIN = Side(style="thin", color="D6DCE5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hex(value: str | None, default: str | None = None) -> str | None:
    if not value:
        return default
    value = value.strip().lstrip("#")
    if len(value) == 3:
        value = "".join(ch * 2 for ch in value)
    return value.upper() if len(value) == 6 else default


def _coerce(value: Any) -> Any:
    """Numbers should land in Excel as numbers, not text."""
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text or not NUMERIC_RE.match(text):
        return value
    cleaned = text.replace(",", "")
    try:
        if cleaned.endswith("%"):
            return float(cleaned[:-1]) / 100
        return int(cleaned) if "." not in cleaned else float(cleaned)
    except ValueError:
        return value


def _style_cell(cell: Any, style: dict[str, Any] | None, *, header: bool = False) -> None:
    style = style or {}
    color = _hex(style.get("color"), "FFFFFF" if header else "1F2937")
    cell.font = Font(
        name=style.get("fontFamily") or "Calibri",
        size=float(style.get("fontSize") or (11 if header else 11)),
        bold=bool(style.get("bold", header)),
        italic=bool(style.get("italic")),
        underline="single" if style.get("underline") else None,
        color=color,
    )
    fill = _hex(style.get("bg"))
    if header and not fill:
        cell.fill = HEADER_FILL
    elif fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(
        horizontal=style.get("align") or ("center" if header else None),
        vertical="center",
        wrap_text=bool(style.get("wrap")),
    )
    cell.border = BORDER


def grid_to_xlsx(grid: dict[str, Any], title: str = "Sheet1") -> bytes:
    headers = [str(h) for h in grid.get("headers") or []]
    rows = grid.get("rows") or []
    cell_styles: dict[str, Any] = grid.get("cellStyles") or {}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = re.sub(r"[\\/*?:\[\]]", "-", (title or "Sheet1"))[:31] or "Sheet1"

    base_header_style = grid.get("headerStyle") or {}
    for column, header in enumerate(headers, start=1):
        # Per-column overrides are stored by the editor under "h-<column index>".
        override = cell_styles.get(f"h-{column - 1}") or {}
        _style_cell(
            sheet.cell(row=1, column=column, value=header),
            {**base_header_style, **override},
            header=True,
        )

    for row_index, row in enumerate(rows):
        if not any(str(cell).strip() for cell in row if cell is not None):
            continue
        for col_index, value in enumerate(row):
            cell = sheet.cell(row=row_index + 2, column=col_index + 1, value=_coerce(value))
            _style_cell(cell, cell_styles.get(f"{row_index}-{col_index}"))

    widths = [len(h) for h in headers]
    for row in rows:
        for index, value in enumerate(row):
            if index < len(widths):
                widths[index] = max(widths[index], len(str(value or "")))
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(max(width + 4, 12), 60)

    sheet.freeze_panes = "A2"
    if headers and sheet.max_row > 1:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def grid_to_csv(grid: dict[str, Any]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer)
    headers = grid.get("headers") or []
    if headers:
        writer.writerow(headers)
    for row in grid.get("rows") or []:
        if any(str(cell).strip() for cell in row if cell is not None):
            writer.writerow(["" if cell is None else cell for cell in row])
    return buffer.getvalue().encode("utf-8-sig")
